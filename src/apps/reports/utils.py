import io
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

def export_to_excel(data, filename):
    """Xuất danh sách dict ra file Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo"
    
    if not data:
        return HttpResponse("Không có dữ liệu", status=404)
    
    # Lấy headers từ keys của dict đầu tiên
    headers = list(data[0].keys())
    
    # Định dạng header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Đổ dữ liệu
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, key in enumerate(headers, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row_data.get(key, ''))
    
    # Tự động điều chỉnh độ rộng cột
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Lưu vào response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


def export_to_pdf(data, filename):
    """Xuất danh sách dict ra file PDF (đơn giản)"""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    if not data:
        return HttpResponse("Không có dữ liệu", status=404)
    
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    doc = SimpleDocTemplate(response, pagesize=A4)
    elements = []
    
    # Tiêu đề
    styles = getSampleStyleSheet()
    title_style = styles['Title']
    elements.append(Paragraph("Báo cáo tổng hợp", title_style))
    elements.append(Spacer(1, 12))
    
    # Bảng dữ liệu
    headers = list(data[0].keys())
    table_data = [headers]
    for item in data:
        row = [str(item.get(h, '')) for h in headers]
        table_data.append(row)
    
    # Tạo bảng
    table = Table(table_data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(table)
    
    doc.build(elements)
    return response